import os
import sys
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from modules import find_excel_files, load_hs_codes, normalize_code

logger = logging.getLogger("exportR." + __name__)

HS_CODE_FILE = "resources/hs_code.json"
HIGHLIGHT_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


def highlight_cargoes_file(file_path, hs_codes):
    wb = load_workbook(file_path)
    highlighted = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=5, max_col=5):
            cell = row[0]
            if cell.value is None:
                continue
            if normalize_code(cell.value) in hs_codes:
                cell.fill = HIGHLIGHT_FILL
                highlighted += 1

    if highlighted:
        wb.save(file_path)

    return highlighted


def run_hs_check(root_folder, status_label=None, row_callback=None):
    hs_codes = load_hs_codes(get_resource_path(HS_CODE_FILE))
    files = find_excel_files(str(root_folder), pattern=r"^CARGOES_LIST")

    processed = 0
    failed = 0
    total_highlighted = 0

    for f in files:
        if status_label:
            status_label.config(text=f"📄 Checking: {f}")
            status_label.update_idletasks()
        try:
            highlighted = highlight_cargoes_file(f, hs_codes)
            total_highlighted += highlighted
            processed += 1
            logger.info(f"Checked '{f}': highlighted {highlighted} cell(s)")
            if highlighted and row_callback:
                row_callback(str(f), highlighted)
        except Exception:
            failed += 1
            logger.exception(f"Failed to process file: {f}")

    if failed:
        summary = (
            f"⚠️ Done. Checked {processed} file(s), highlighted {total_highlighted} cell(s), "
            f"{failed} file(s) lỗi — xem log."
        )
        if status_label:
            status_label.config(text=summary, fg="red")
    else:
        summary = f"✅ Done. Checked {processed} file(s), highlighted {total_highlighted} cell(s)."
        if status_label:
            status_label.config(text=summary)
    logger.info(summary)


def get_resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)

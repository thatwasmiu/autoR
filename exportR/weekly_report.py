import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import re

TEMPLATE_FILE = "resources/template_weekly.xlsx"
OUTPUT_FILE = "output.xlsx"


def find_header_row(sheet, keyword="STT"):
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip().upper() == keyword:
                return cell.row
    return None

COLUMN_MAPPING = {
    "STT NVL": "nvlCode",
    "申报日期": "date",
    "BILL": "bill",
    "线": "routeType",
    "发票号": "invoice",
    "时间": "time"
}

def normalize(text):
    if not text:
        return ""
    
    parts = str(text).splitlines()  # split by \n
    parts = [p.strip().upper() for p in parts if p.strip()]
    
    # return all parts joined OR last part
    return parts[-1] if parts else ""

def extract_data_from_sheet(sheet, method):
    header_row_idx = find_header_row(sheet)
    if not header_row_idx:
        return []

    # Read header row
    header_row = list(sheet.iter_rows(
        min_row=header_row_idx,
        max_row=header_row_idx,
        values_only=True
    ))[0]

    # Map normalized column name -> index
    col_map = {}
    for idx, col_name in enumerate(header_row):
        norm = normalize(col_name)
        if norm:
            col_map[norm] = idx

    # Normalize mapping keys
    normalized_mapping = {
        normalize(k): v for k, v in COLUMN_MAPPING.items()
    }

    # Check required columns exist
    missing = [k for k in normalized_mapping if k not in col_map]
    if missing:
        print(f"⚠️ Missing columns {missing} in sheet {sheet.title}")
        return []

    data = []

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        row_dict = {"method": method}
        for norm_col, code_key in normalized_mapping.items():
            row_dict[code_key] = row[col_map[norm_col]]

        data.append(row_dict)

    return data

def extract_data_from_file(filepath, status_label):
    wb = load_workbook(filepath, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        sheet_data = extract_data_from_sheet(sheet, sheet_name)

        if sheet_data:
            all_data.extend(sheet_data)
        else:
            status_label.config(text="Cannot find the label STT")

    return all_data

OUTPUT_COLUMN_MAP = {
    "date": 2,
    "nvlCode": 3,
    "bill": 4,
    "invoice": 5,
    "time": 10, 
    "routeType": 11,
    "method": 13
}

KEY_FIELDS = ["nvlCode", "bill", "invoice"]


def build_existing_index(sheet, start_row):
    """
    Build a map:
    (nvlCode, bill, invoice) -> row_number
    """
    index = {}

    for row in sheet.iter_rows(min_row=start_row, values_only=False):
        row_idx = row[0].row

        key = []
        for field in KEY_FIELDS:
            col_idx = OUTPUT_COLUMN_MAP[field]
            val = sheet.cell(row=row_idx, column=col_idx).value
            key.append(str(val).strip() if val else "")

        key_tuple = tuple(key)

        if any(key_tuple):  # ignore empty rows
            index[key_tuple] = row_idx

    return index

from datetime import datetime

def parse_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value

    # try common formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except:
            continue

    return None  # fallback

def parse_folder_date(name, year):
    match = re.match(r"(\d{1,2})\.(\d{1,2})", name)
    if not match:
        return None
    d, m = map(int, match.groups())
    try:
        return datetime(year, m, d)
    except:
        return None

def create_weekly_report(root_folder, from_date, to_date, status_label):
    all_data = []

    # user selected range
    from_dt = datetime.strptime(from_date, "%d/%m/%Y")
    to_dt = datetime.strptime(to_date, "%d/%m/%Y")

    current_year = from_dt.year  # assume same year

    for root, dirs, files in os.walk(root_folder):
        data = None
        for d in dirs:
            # match folder like "21.04"
            folder_date = parse_folder_date(d, current_year)
            if folder_date:
                # ✅ filter by range
                if from_dt <= folder_date <= to_dt:
                    folder_path = os.path.join(root, d)

                    for file in os.listdir(folder_path):
                        if file.endswith(".xlsx") and "BC_" in file:
                            full_path = os.path.join(folder_path, file)
                            status_label.config(text=f"📄 Processing: {full_path}")
                            data = extract_data_from_file(full_path, status_label)
                            all_data.extend(data)

    if not all_data:
        status_label.config(text="⚠️ No data found")
        return

    wb = load_workbook(get_resource_path(TEMPLATE_FILE))
    sheet = wb.active
    now = datetime.today()
    sheet.title = f"T{now.month}.{now.year}"

    # 🔍 find header row
    header_row = find_header_row(sheet)
    if not header_row:
        raise Exception("❌ Cannot find header row (STT)")

    data_start_row = header_row + 1
    status_label.config(text=f"Start printing!!!")
    # 🔧 build existing index
    existing_index = build_existing_index(sheet, data_start_row)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    font_tnr = Font(name="Times New Roman", size=10) 

    label_row = sheet.max_row
    append_row = label_row + 1
    updated = 0
    inserted = 0

    for row_dict in all_data:
        key = tuple(str(row_dict.get(f, "")).strip() for f in KEY_FIELDS)
        row = None
        if key in existing_index:
            # 🔁 UPDATE existing row
            row = existing_index[key]
            indexCell = sheet.cell(row=row, column=1)
            indexCell.value = row - label_row
            for k, col_idx in OUTPUT_COLUMN_MAP.items():
                if k not in KEY_FIELDS:  # don't overwrite key fields
                    value = row_dict.get(k)

                    cell = sheet.cell(row=row, column=col_idx)

                    if k == "date":
                        dt = parse_date(value)
                        cell.value = dt
                        if dt:
                            cell.number_format = "D/M/YYYY"
                    if k == "method" and value != None and value.upper() == "TRUCK":      
                        cell.value = ""        
                    else:
                        cell.value = value
            updated += 1

        else:
            # ➕ INSERT new row
            row = append_row
            indexCell = sheet.cell(row=row, column=1)
            indexCell.value = row - label_row
            for k, col_idx in OUTPUT_COLUMN_MAP.items():
                value = row_dict.get(k)

                cell = sheet.cell(row=append_row, column=col_idx)

                if k == "date":
                    dt = parse_date(value)
                    cell.value = dt
                    if dt:
                        cell.number_format = "D/M/YYYY"
                if k == "method" and value != None and value.upper() == "TRUCK":      
                    cell.value = ""
                else:
                    cell.value = value

            existing_index[key] = append_row
            append_row += 1
            inserted += 1       

        for col in range(1, 14):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.font = font_tnr   # ✅ apply Times New Roman
    now = datetime.now()
    timestamp = f"T{now.isocalendar().week:02d}_{now.strftime('%Y')}"
    output_file = root_folder / f"BC_{timestamp}_{now.strftime('%H%M%S')}.xlsx"
    status_label.config(text=f"✅ Done! Saved: {str(output_file)}")
    wb.save(output_file)
    os.startfile(output_file)

def get_resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)    
from openpyxl import load_workbook
import re
import os

def find_values(filename, patterns):
    wb = load_workbook(filename, data_only=True)

    compiled = {}

    for k, cfg in patterns.items():
        kw = cfg.get("kw", "").lower()
        p = cfg.get("regex")
        files = cfg.get("files")
        sheets = cfg.get("sheets")

        compiled[k] = {
            "keyword": kw,
            "pattern": re.compile(p, re.IGNORECASE) if p else None,
            "files": files,
            "sheets": sheets,
            "merged": cfg.get("merged", False)
        }

    results = {k: None for k in patterns}
    # if "合同_发票_箱单" in os.path.basename(filename):
    #     print(kw)

    # if kw == 'no.':
    #     print(compiled)

    for ws in wb.worksheets:
        sheet_name = ws.title

        # A merged cell only stores its value on the top-left (anchor) cell;
        # every other cell in the range reads back as None from openpyxl.
        # Map every coordinate in each merged range to that range's value so
        # the column scan below can't land on a "None" cell inside a merge
        # and skip past the real value.
        merged_lookup = {}
        for merged_range in ws.merged_cells.ranges:
            anchor_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if anchor_value is None:
                continue
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_lookup[(r, c)] = anchor_value

        for key, cfg in compiled.items():
            # skip if already found
            if results[key] is not None:
                continue 

            # check filename condition
            if cfg["files"] is not None:
                base = os.path.splitext(os.path.basename(filename))[0]

                # if "合同_发票_箱单" in os.path.basename(filename) and cfg["keyword"]=="no.":
                #     print("BASE RAW:", base)
                #     print("BASE REPR:", repr(base))
                #     for p in cfg["files"]:
                #         print("PATTERN:", p)
                #         print("MATCH:", re.search(p, base, re.IGNORECASE)) 

                if not base or not any(
                    re.search(p, base, re.IGNORECASE)
                    for p in cfg["files"]
                ):
                    continue

            # check sheet condition
            if cfg["sheets"] is not None:
                if sheet_name not in cfg["sheets"]:
                    continue

            keyword = cfg["keyword"]
            pattern = cfg["pattern"]

            # if (keyword == 'no.'):
                # print(cfg["sheets"])

            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value:
                        continue

                    text = str(cell.value).lower()
                    # print(text)

                    if keyword in text:
                        extra_cols = 30 if cfg["merged"] else 8 
                        for col in range(cell.column + 1, cell.column + extra_cols):
                            val = ws.cell(row=cell.row, column=col).value
                            if val is None and cfg["merged"]:
                                val = merged_lookup.get((cell.row, col))
                            if not val:
                                continue

                            if pattern:
                                m = pattern.search(str(val))
                                if m:
                                    results[key] = m.group(0).strip()
                                    break
                            else:
                                results[key] = str(val).strip()
                                break

                if results[key] is not None:
                    break

        # stop early if all found
        if all(results.values()):
            break
    # if "合同_发票_箱单" in os.path.basename(filename):        
    #     print(results)
    return results



def get_workbook(file_path):
    wb = load_workbook(file_path, data_only=True)
    return wb
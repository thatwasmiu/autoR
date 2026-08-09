from copy import copy
from openpyxl import load_workbook
from datetime import datetime, time
from openpyxl.styles import Border, Side
from openpyxl.styles import Font

switch = {
    '1': "Xanh",
    '2': "Vàng",
    '3': "Đỏ"
}

def write_daily_report(template, grouped):
    # print(grouped)
    wb = load_workbook(template)

    template_ws = wb.active  # original styled sheet

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    font_tnr = Font(name="Times New Roman", size=10)

    for method, items in grouped.items():

        valid_items = [data for data in items if data.get("declareCode") is not None]
        if not valid_items:
            continue

        ws = wb.copy_worksheet(template_ws)
        ws.title = str(method).upper()

        # optional: clear existing data rows (keep header)
        # ws.delete_rows(2, ws.max_row)

        is_e15 = "E15" in ws.title
        if is_e15:
            route_header = ws.cell(row=3, column=5)  # "线"
            internal_header = ws.cell(row=3, column=6)
            internal_header.value = "Số tờ khai xuất"
            internal_header.font = copy(route_header.font)
            internal_header.fill = copy(route_header.fill)
            internal_header.border = copy(route_header.border)
            internal_header.alignment = copy(route_header.alignment)
            ws.column_dimensions['F'].width = 20

        for i, data in enumerate(valid_items, start=1):
            date_val = None
            try:
                date_str = data.get("date")
                if date_str:
                    date_val = datetime.strptime(date_str, "%d/%m/%Y %H:%M:%S")
            except (ValueError, TypeError):
                date_val = None

            time_str = None
            if date_val:
                time_str = date_val.strftime("%I:%M %p")

            route_code = data.get("routeType")
            route_label = switch.get(route_code, "")
            if route_code in ("1", "2") and date_val and date_val.time() > time(17, 0):
                route_label = f"{route_label} OT"

            row_values = [
                i,
                # data.get("month"),
                data.get("nvlCode"),
                # "DONE" if data.get("isDone") else "PENDING",
                # data.get("formCode"),
                date_val,
                # data.get("bill"),
                # "HQ TELECOM",
                data.get("declareCode"),
                route_label,
                # data.get("typeCode"),
                # data.get("term"),
                # data.get("invoice"),
                # data.get("tms"),
                # time_str,
            ]
            if is_e15:
                row_values.append(data.get("internalCode"))

            ws.append(row_values)

            row = ws.max_row

            if date_val:
                ws.cell(row=row, column=3).number_format = "dd/mm/yyyy"

            last_col = 6 if is_e15 else 5
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.font = font_tnr   # ✅ apply Times New Roman
    wb.remove(template_ws)
    return wb
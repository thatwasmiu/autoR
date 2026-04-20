import os
from openpyxl import load_workbook

ROOT_FOLDER = r"C:\Users\datnt4\Documents\week"
TEMPLATE_FILE = "template_weekly.xlsx"
OUTPUT_FILE = "output.xlsx"


def find_header_row(sheet, keyword="STT"):
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip().upper() == keyword:
                return cell.row
    return None

COLUMN_MAPPING = {
    "STT NVL": "nvlCode",
    "申报日期": "date",
    "BILL": "bill",
    "线": "routeType",
    "发票号": "invoice",
    "时间": "time"
}

def normalize(text):
    if not text:
        return ""
    
    parts = str(text).splitlines()  # split by \n
    parts = [p.strip().upper() for p in parts if p.strip()]
    
    # return all parts joined OR last part
    return parts[-1] if parts else ""

def extract_data_from_sheet(sheet):
    header_row_idx = find_header_row(sheet)
    if not header_row_idx:
        return []

    # Read header row
    header_row = list(sheet.iter_rows(
        min_row=header_row_idx,
        max_row=header_row_idx,
        values_only=True
    ))[0]

    # Map normalized column name -> index
    col_map = {}
    for idx, col_name in enumerate(header_row):
        norm = normalize(col_name)
        if norm:
            col_map[norm] = idx

    # Normalize mapping keys
    normalized_mapping = {
        normalize(k): v for k, v in COLUMN_MAPPING.items()
    }

    # Check required columns exist
    missing = [k for k in normalized_mapping if k not in col_map]
    if missing:
        print(f"⚠️ Missing columns {missing} in sheet {sheet.title}")
        return []

    data = []

    # Read data rows
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue

        row_dict = {}
        for norm_col, code_key in normalized_mapping.items():
            row_dict[code_key] = row[col_map[norm_col]]

        data.append(row_dict)

    return data

def extract_data_from_file(filepath):
    wb = load_workbook(filepath, data_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"   ➜ Sheet: {sheet_name}")

        sheet_data = extract_data_from_sheet(sheet)

        if sheet_data:
            all_data.extend(sheet_data)
        else:
            print(f"      ⚠️ No STT found")

    return all_data

OUTPUT_COLUMN_MAP = {
    "date": 2,
    "nvlCode": 3,
    "bill": 4,
    "invoice": 5,
    "routeType": 8
}

def main():
    all_data = []

    for root, dirs, files in os.walk(ROOT_FOLDER):
        for file in files:
            if file.endswith(".xlsx") and "BC_" in file:
                full_path = os.path.join(root, file)
                print(f"📄 Processing: {full_path}")

                data = extract_data_from_file(full_path)
                all_data.extend(data)

    if not all_data:
        print("⚠️ No data found")
        return

    wb = load_workbook(TEMPLATE_FILE)  # use existing file
    sheet = wb.active

    # 👉 append to last row
    start_row = sheet.max_row + 1

    for row_idx, row_dict in enumerate(all_data, start=start_row):
        for key, col_idx in OUTPUT_COLUMN_MAP.items():
            sheet.cell(
                row=row_idx,
                column=col_idx,
                value=row_dict.get(key)
            )

    wb.save(OUTPUT_FILE)
    print(f"✅ Appended {len(all_data)} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
import re
from collections import Counter
from datetime import datetime
from openpyxl import load_workbook
import os
from .db import DeclareForm

patterns = {
    "declare_code": {
        "kw": "Số tờ khai",
        "regex": None,
        "files": [r"ToKhai.*_\d+"],
        "sheets": None
    },
    "type_code": {
        "kw": "Mã loại hình",
        "regex": r'[A-Z]+\d+\s+\d+\s+\[\s*\d+\s*\]',
        "files": [r"ToKhai.*_\d+"],
        "sheets": None
    },
    "route_type": {
        "kw": "Mã phân loại kiểm tra",
        "regex": r"\d",
        "files": [r"ToKhai.*_\d+"],
        "sheets": None
    },
    "term": {
        "kw": "Tổng trị giá hóa đơn",
        "regex": None,
        "files": [r"ToKhai.*_\d+"],
        "sheets": None
    },
    "date": {
        "kw": "Ngày đăng ký",
        "regex": None,
        "files": [r"ToKhai.*_\d+"],
        "sheets": None
    },
    "invoice": {
        "kw": "NO.",
        "regex": None,
        "files": [r"合同_发票_箱单.*"],
        "sheets": ['INVForm']
    },
    "method": {
        "kw": "Ship By:",
        "regex": None,
        "files": [r"合同_发票_箱单.*"],
        "sheets": None
    },
}

route_switch = {
    '1': "Xanh",
    '2': "Vàng",
    '3': "Đỏ"
}

method_switch = {
    '1': "AIR",
    '2': "SEA",
    '3': "SEA",
    '4': "TRUCK"
}

def get_data(daily_invoice_folder):
    # folder = r"C:\Users\datnt4\Documents\06.04\1. NVL - 9365 - E20260403058  - LENOVOVN20260406003 - 6.4.2026 - GC - 2PK - E11- TRUCK"
    folder = str(daily_invoice_folder)
    form_code = get_form_code(daily_invoice_folder.name)
    # print(nvl_code, bill, invoice)
    # return

    files = find_excel_files(folder)

    declare_codes = []
    type_codes = []
    route_types = []
    terms = []
    dates = []
    tmses = []
    invoices = []
    methods = []
    for f in files:
        try:
            values = find_values(f, patterns)
        except Exception as e:
            print(f"Error processing files {f}: {e}")
            continue

        if values["declare_code"]:
            declare_codes.append(values["declare_code"].strip())
        if values["type_code"]:
            type_codes.append(values["type_code"].strip())
        if values["route_type"]:
            route_types.append(values["route_type"].strip())
        if values["term"]:
            terms.append(values["term"].strip())
        if values["date"]:
            dates.append(values["date"].strip())
        if values["invoice"]:
            invoices.append(values["invoice"].strip())
        if values["method"]:
            methods.append(values["method"].strip().upper())

        if "合同_发票_箱单" in f:
            tms = get_tms_code(f)
            if tms:
                tmses.append(tms)   
    
    declare_code = pick_value(declare_codes, folder, r"[A-Za-z0-9]+")
    type_code = pick_value(type_codes, folder, r"[A-Za-z0-9]+")
    route_type = pick_value(route_types, folder, r"[A-Za-z0-9]+")
    term = pick_value(terms, folder, r'^\s*[^-]+\s*-\s*([^-]+)\s*-', 1)
    date = pick_value(dates, folder, r'\b(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4} ([01][0-9]|2[0-3]):[0-5][0-9]:[0-5][0-9]\b')
    tms = pick_value(tmses)
    invoice = pick_value(invoices, folder)

    nvl_code, bill = get_codes(daily_invoice_folder.name, invoice)
    month = None
    time = None
    date_str = None

    if date:
        try:
            parsed_date = datetime.strptime(date, "%d/%m/%Y %H:%M:%S")
            month = parsed_date.month
            time = parsed_date.strftime("%I:%M %p")
            date_str = parsed_date.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            print("Error date:", date)

    method = pick_value(type_codes, folder, r'(\d+)\s*(?=\[)', 1)
    method_str = pick_value(methods, folder)
    if method:
        method = method_switch.get(method.strip(), method_str)  # map to AIR/SEA/TRUCK or keep original if not in switch

    # print(method)
    return DeclareForm(
        nvl_code=nvl_code,
        bill=bill,
        invoice=invoice,
        declare_code=declare_code,
        type_code=type_code,
        route_type=route_switch.get(route_type, ""),
        term=term,
        date=date_str,
        month=month,
        tms=tms,
        form_code=form_code,
        method=method,
        time=time
    )

def get_form_code(folder_name): 
    parts = [p.strip() for p in folder_name.split("-")]

    for p in parts:
        if p == "GC" or p == "CX":
            return p
    return "GC"

def get_tms_code(file_name):
    m = re.search(r'(?<=合同_发票_箱单_)I\d+', file_name)

    if m:
        return m.group(0).strip()
    return None

def pick_value(values, folder=None, format_regex=None, group_index=0):
    if not values:
        return None

    # NEW: transform values using regex group(0)
    if format_regex:
        pattern = re.compile(format_regex)
        new_values = []

        for v in values:
            m = pattern.search(str(v))
            if m:
                new_values.append(m.group(group_index))

        values = new_values

    if not values:
        return None

    # keep only normal values
    # values = [v for v in values if is_normal(v)]

    counter = Counter(values)
    most_common = counter.most_common()

    # if top count > 1 → return most frequent
    if most_common[0][1] > 1:
        return most_common[0][0]

    # Find all values that appear in 
    if folder:
        matched = [v for v in values if v in folder]
        if matched:
            return ";".join(matched)

    # fallback → join all
    return ";".join(values)

# Filter out special characters (keep letters, numbers, maybe _)
def is_normal(v):
    return re.fullmatch(r"[A-Za-z0-9]+", v) is not None    


def find_values(filename, patterns):
    wb = load_workbook(filename, data_only=True)

    compiled = {}

    for k, cfg in patterns.items():
        kw = cfg.get("kw", "").lower()
        p = cfg.get("regex")
        files = cfg.get("files")
        sheets = cfg.get("sheets")

        compiled[k] = {
            "keyword": kw,
            "pattern": re.compile(p, re.IGNORECASE) if p else None,
            "files": files,
            "sheets": sheets
        }

    results = {k: None for k in patterns}
    # if "合同_发票_箱单" in os.path.basename(filename):
    #     print(kw)

    # if kw == 'no.':
    #     print(compiled)

    for ws in wb.worksheets:
        sheet_name = ws.title

        for key, cfg in compiled.items():
            # skip if already found
            if results[key] is not None:
                continue 

            # check filename condition
            if cfg["files"] is not None:
                base = os.path.splitext(os.path.basename(filename))[0]

                # if "合同_发票_箱单" in os.path.basename(filename) and cfg["keyword"]=="no.":
                #     print("BASE RAW:", base)
                #     print("BASE REPR:", repr(base))
                #     for p in cfg["files"]:
                #         print("PATTERN:", p)
                #         print("MATCH:", re.search(p, base, re.IGNORECASE)) 

                if not base or not any(
                    re.search(p, base, re.IGNORECASE)
                    for p in cfg["files"]
                ):
                    continue

            # check sheet condition
            if cfg["sheets"] is not None:
                if sheet_name not in cfg["sheets"]:
                    continue

            keyword = cfg["keyword"]
            pattern = cfg["pattern"]

            # if (keyword == 'no.'):
                # print(cfg["sheets"])

            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value:
                        continue

                    text = str(cell.value).lower()
                    # print(text)

                    if keyword in text:
                        for col in range(cell.column + 1, cell.column + 8):
                            val = ws.cell(row=cell.row, column=col).value
                            if not val:
                                continue

                            if pattern:
                                m = pattern.search(str(val))
                                if m:
                                    results[key] = m.group(0).strip()
                                    break
                            else:
                                results[key] = str(val).strip()
                                break

                if results[key] is not None:
                    break

        # stop early if all found
        if all(results.values()):
            break
    # if "合同_发票_箱单" in os.path.basename(filename):        
    #     print(results)
    return results

def get_workbook(file_path):
    wb = load_workbook(file_path, data_only=True)
    return wb

def find_excel_files(folder, pattern=None):
    files = []
    for root, dirs, filenames in os.walk(folder):
        for f in filenames:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                if pattern and not re.search(pattern, f):
                    continue
                files.append(os.path.join(root, f))
    return files

def get_codes(text, invoice):

    pattern = rf"-\s*{re.escape(invoice)}"
    text = re.split(pattern, text, maxsplit=1)[0]
    # remove spaces around "-"
    parts = [p.strip() for p in re.split(r'-', text)]

    # find NVL index
    try:
        i = next(i for i, p in enumerate(parts) if "NVL" in p)
    except StopIteration:
        return None, None

    # first
    if i + 1 >= len(parts):
        return None, None
#     first = f"{parts[i]}-{parts[i+1]}"
    first = f"NVL - {parts[i+1]}"

    # second: aaa or aaa-01
    if i + 2 >= len(parts):
        return first, None

    second = "-".join(parts[i+2:])

    return first, second
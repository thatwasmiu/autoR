import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from datetime import datetime
import re

TEMPLATE_FILE = "resources/template_weekly.xlsx"
OUTPUT_FILE = "output.xlsx"

def create_weekly_report(root_folder, all_data, status_label):

    if not all_data:
        status_label.config(text="⚠️ No data found")
        return
    
    print(all_data)

    wb = load_workbook(get_resource_path(TEMPLATE_FILE))
    sheet = wb.active
    now = datetime.today()
    sheet.title = f"T{now.month}.{now.year}"

    # 🔍 find header row
    
    status_label.config(text=f"Start printing!!!")

    header_row = find_header_row(sheet)
    row = header_row + 1

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    font_tnr = Font(name="Times New Roman", size=10) 
    
    i = 1
    for data in all_data:
        date_val = None
        try:
            date_str = data.get("date")
            if date_str:
                date_val = datetime.strptime(date_str, "%d/%m/%Y %H:%M:%S")
        except (ValueError, TypeError):
            date_val = None

        sheet.append([
                    i,
                    date_val,
                    data.get("nvl_code"),
                    data.get("bill"),
                    data.get("invoice"),
                    data.get("mail_time"),
                    data.get("tms_time"),
                    data.get("draft_time"),
                    data.get("tk_time"),
                    data.get("official_time"),
                    data.get("passed_time"),
                    data.get("route_type"),
                    data.get("method") if (data.get("method") or "").lower() != "truck" else "",
                ])
        if date_val:
            sheet.cell(row=row, column=2).number_format = "D/M/YYYY"

        for col in range(1, 14):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            cell.font = font_tnr   # ✅ apply Times New Roman
        i += 1
        row += 1
    now = datetime.now()
    timestamp = f"T{now.isocalendar().week:02d}_{now.strftime('%Y')}"
    output_file = root_folder / f"BC_{timestamp}_{now.strftime('%H%M%S')}.xlsx"
    status_label.config(text=f"✅ Done! Saved: {str(output_file)}")
    wb.save(output_file)
    os.startfile(output_file)

def parse_date(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value

    # try common formats
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except:
            continue

    return None  # fallback

def find_header_row(sheet, keyword="STT"):
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip().upper() == keyword:
                return cell.row
    return None

def get_resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)    
from openpyxl import load_workbook
from datetime import datetime, time
from openpyxl.styles import Border, Side
from openpyxl.styles import Font

switch = {
    '1': "Xanh",
    '2': "Vàng",
    '3': "Đỏ"
}

def write_daily_report(template, grouped):
    # print(grouped)
    wb = load_workbook(template)

    template_ws = wb.active  # original styled sheet

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    font_tnr = Font(name="Times New Roman", size=10)

    first = True

    for method, items in grouped.items():

        # ✅ Use template sheet for first group, copy for others
        if first:
            ws = wb.copy_worksheet(template_ws)
            ws.title = str(method).upper()
            first = False
        else:
            ws = wb.copy_worksheet(template_ws)
            ws.title = str(method).upper()

        # optional: clear existing data rows (keep header)
        # ws.delete_rows(2, ws.max_row)

        for i, data in enumerate(items, start=1):
            date_val = None
            try:
                date_str = data.get("date")
                if date_str:
                    date_val = datetime.strptime(date_str, "%d/%m/%Y %H:%M:%S")
            except (ValueError, TypeError):
                date_val = None

            time_str = None
            if date_val:
                time_str = date_val.strftime("%I:%M %p")

            ws.append([
                i,
                data.get("month"),
                data.get("nvlCode"),
                None,
                data.get("formCode"),
                date_val,
                data.get("bill"),
                "HQ TELECOM",
                data.get("declareCode"),
                switch.get(data.get("routeType"), ""),
                data.get("typeCode"),
                data.get("term"),
                data.get("invoice"),
                data.get("tms"),
                time_str,
            ])

            row = ws.max_row

            if date_val:
                ws.cell(row=row, column=6).number_format = "D/M/YYYY"

            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.font = font_tnr   # ✅ apply Times New Roman
    wb.remove(template_ws)
    return wb